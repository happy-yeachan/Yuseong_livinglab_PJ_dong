from data_control import *
from openpyxl import Workbook
import datetime
from PyQt5.QtWidgets import QMessageBox

def show_current(screen):
    screen.reset_button_styles()
    screen.current_button.setStyleSheet('background-color: lightblue')
    screen.label.setText('참전 명예 수당 지급자 현황')
    screen.table.setColumnCount(13)
    screen.table.setHorizontalHeaderLabels([
        '동', '입력날짜', '보훈번호', '성명', '주민번호', '주소',
        '입금유형', '은행명', '예금주', '계좌번호', '신규 사유', '전입일', '비고'
    ])
    rows = get_data("Honor_of_War_Current")

    screen.load_data(rows, 'now')


def export_to_excel_Now():
    # 워크북 생성 (각 시트가 추가될 때까지 유지됨)
    wb = Workbook()
    ws = wb.active
    ws.title = "현황"

    # '현황' 시트 헤더 추가
    headers = [
        '동', '등록일', '보훈번호', '성명', '주민번호',
        '주소', '입금유형', '은행명', '예금주', '계좌번호',
        '신규사유', '전입일', '비고'
    ]
    ws.append(headers)

    # 데이터를 가져오는 함수 호출
    rows = get_data("Honor_of_War_Current")

    # 데이터 추가
    for row in rows:
        ws.append(row)

    # 모든 열의 크기 자동 조정
    adjust_column_width(ws)

    # '신규자' 및 '중지자' 시트 생성
    export_to_excel_New(wb)
    export_to_excel_Stop(wb)

    # 파일 이름 생성 (동명 + 현재 날짜 기반)
    dong_name = rows[0][0]  # 첫 번째 행의 첫 번째 필드에서 동명 가져오기
    current_date = datetime.datetime.now().strftime("%Y%m")
    file_name = f"참전유공자_{current_date}_{dong_name}.xlsx"

    # 엑셀 파일 저장
    wb.save(file_name)

    # 저장 완료 메시지 박스 표시
    QMessageBox.information(None, "엑셀추출", f"파일명: {file_name} \n성공적으로 저장되었습니다!")

def export_to_excel_New(wb):
    # '신규자' 시트 추가
    ws = wb.create_sheet(title="신규자")

    # '신규자' 시트 헤더 추가
    headers = [
        '동', '등록일', '보훈번호', '성명', '주민번호',
        '주소', '입금유형', '은행명', '예금주', '계좌번호',
        '사유', '전입일', '비고'
    ]
    ws.append(headers)

    # 데이터를 가져오는 함수 호출
    rows = get_data("Honor_of_War_New")

    # 데이터 추가
    for row in rows:
        ws.append(row)

    # 모든 열의 크기 자동 조정
    adjust_column_width(ws)

def export_to_excel_Stop(wb):
    # '중지자' 시트 추가
    ws = wb.create_sheet(title="중지자")

    # '중지자' 시트 헤더 추가
    headers = [
        '동', '등록일', '보훈번호', '성명', '주민번호',
        '주소', '전입일', '중단사유', '사유일시', '비고'
    ]
    ws.append(headers)

    # 데이터를 가져오는 함수 호출
    rows = get_data("Honor_of_War_Stop")

    # 데이터 추가
    for row in rows:
        ws.append(row[:10])

    # 모든 열의 크기 자동 조정
    adjust_column_width(ws)

def adjust_column_width(ws):
    # 모든 열의 크기 자동 조정 함수
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        # 열 너비 설정
        adjusted_width = (max_length + 3)
        ws.column_dimensions[column_letter].width = adjusted_width